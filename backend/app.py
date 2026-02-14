import os
import json
import logging
from datetime import datetime, time
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from itsdangerous import URLSafeSerializer, BadSignature
import io

# Set up logging
logging.basicConfig(level=logging.DEBUG)

# Determine if running in Docker container or local development
if os.path.exists('/app'):  # Docker container
    template_folder = 'frontend/templates'
    static_folder = 'frontend/static'
    data_dir = 'database/data'
else:  # Local development (via main.py)
    template_folder = '../frontend/templates'
    static_folder = '../frontend/static'
    data_dir = '../database/data'

app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
app.secret_key = os.environ.get("SESSION_SECRET", "dev_secret_key_change_in_production")
token_serializer = URLSafeSerializer(app.secret_key, salt="attendance-api-token")

# Ensure data directory exists
os.makedirs(data_dir, exist_ok=True)

# Class schedule with time restrictions
CLASS_SCHEDULE = {
    'Mathematics': {'start': time(9, 0), 'end': time(10, 0)},
    'Science': {'start': time(10, 30), 'end': time(11, 30)},
    'English': {'start': time(12, 0), 'end': time(13, 0)},
    'Social Studies': {'start': time(14, 0), 'end': time(15, 0)},
    'Hindi': {'start': time(15, 30), 'end': time(16, 30)}
}

def create_auth_token(payload):
    """Create signed auth token for API consumers."""
    return token_serializer.dumps(payload)

def decode_auth_token(token):
    """Decode signed auth token and return payload."""
    return token_serializer.loads(token)

def extract_token():
    """Extract bearer token from Authorization header or token query param."""
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        return auth_header.split(" ", 1)[1].strip()
    return request.args.get("token", "").strip()

def api_auth_required(role=None):
    """API authentication decorator with optional role restriction."""
    def decorator(func):
        @wraps(func)
        def wrapped(*args, **kwargs):
            token = extract_token()
            if not token:
                return jsonify({"ok": False, "error": "Missing auth token"}), 401
            try:
                payload = decode_auth_token(token)
            except BadSignature:
                return jsonify({"ok": False, "error": "Invalid auth token"}), 401

            if role and payload.get("role") != role:
                return jsonify({"ok": False, "error": "Forbidden"}), 403

            request.auth = payload
            return func(*args, **kwargs)
        return wrapped
    return decorator

def serialize_schedule():
    """Serialize class schedule into JSON-friendly shape."""
    return {
        subject: {
            "start": schedule["start"].strftime("%I:%M %p"),
            "end": schedule["end"].strftime("%I:%M %p"),
        }
        for subject, schedule in CLASS_SCHEDULE.items()
    }

def build_attendance_summary(attendance):
    """Group attendance records by date and subject."""
    attendance_summary = {}
    for record in attendance.values():
        date = record["date"]
        subject = record["subject"]
        if date not in attendance_summary:
            attendance_summary[date] = {}
        if subject not in attendance_summary[date]:
            attendance_summary[date][subject] = []
        attendance_summary[date][subject].append(record)
    return attendance_summary

def build_attendance_workbook(attendance):
    """Create a workbook from attendance records."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Date", "Student Name", "Roll Number", "Subject", "Time Marked"]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for record in attendance.values():
        ws.append([
            record["date"],
            record["student_name"],
            record["roll_number"],
            record["subject"],
            record["time"],
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb

@app.after_request
def add_cors_headers(response):
    """Allow Flutter web dev server to call backend APIs."""
    response.headers["Access-Control-Allow-Origin"] = request.headers.get("Origin", "*")
    response.headers["Access-Control-Allow-Headers"] = "Authorization, Content-Type"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Vary"] = "Origin"
    return response

@app.route("/api/<path:_path>", methods=["OPTIONS"])
def api_options(_path):
    return ("", 204)

def load_students():
    """Load student data from JSON file"""
    try:
        data_file = 'database/data/students.json' if os.path.exists('/app') else '../database/data/students.json'
        with open(data_file, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        # Create initial student data if file doesn't exist
        student_names = [
            'Alex', 'Blake', 'Casey', 'Drew', 'Eli', 'Finn', 'Gray', 'Harper', 'Ivan', 'Jade',
            'Kyle', 'Luna', 'Max', 'Nova', 'Oliver', 'Paris', 'Quinn', 'River', 'Sam', 'Taylor',
            'Uma', 'Victor', 'Wade', 'Xara', 'Yale', 'Zoe', 'Ash', 'Brook', 'Clay', 'Dawn',
            'Ember', 'Ford', 'Grace', 'Heath', 'Iris', 'Juno', 'Knox', 'Lexi', 'Mira', 'Nash'
        ]
        
        students = {}
        for i in range(1, 41):
            roll_number = str(i)
            shortname = student_names[i-1]  # Get unique shortname
            students[shortname.lower()] = {  # Use lowercase shortname as key
                'name': f'{shortname} (Roll {i:02d})',
                'roll_number': roll_number,
                'shortname': shortname.lower(),
                'password': f'pass{i:02d}',
                'class': '10'
            }
        save_students(students)
        return students

def save_students(students):
    """Save student data to JSON file"""
    data_file = 'database/data/students.json' if os.path.exists('/app') else '../database/data/students.json'
    with open(data_file, 'w') as f:
        json.dump(students, f, indent=2)

def load_attendance():
    """Load attendance data from JSON file"""
    try:
        data_file = 'database/data/attendance.json' if os.path.exists('/app') else '../database/data/attendance.json'
        with open(data_file, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

def save_attendance(attendance):
    """Save attendance data to JSON file"""
    data_file = 'database/data/attendance.json' if os.path.exists('/app') else '../database/data/attendance.json'
    with open(data_file, 'w') as f:
        json.dump(attendance, f, indent=2)

def is_class_time(subject):
    """Check if current time falls within the specified class schedule"""
    # For testing purposes, simulate time as 9:10 AM
    simulated_time = time(9, 10)  # 9:10 AM
    current_time = simulated_time
    
    schedule = CLASS_SCHEDULE.get(subject)
    if not schedule:
        logging.debug(f"No schedule found for subject: {subject}")
        return False
    
    is_time = schedule['start'] <= current_time <= schedule['end']
    logging.debug(f"Time check for {subject}: Current={current_time}, Start={schedule['start']}, End={schedule['end']}, Result={is_time}")
    return is_time

def has_marked_attendance_today(roll_number, subject):
    """Check if student has already marked attendance for the subject today"""
    attendance = load_attendance()
    today = datetime.now().strftime('%Y-%m-%d')
    
    attendance_key = f"{roll_number}_{subject}_{today}"
    return attendance_key in attendance

@app.route("/api/health")
def api_health():
    return jsonify({"ok": True, "service": "attendance-backend"})

@app.route("/api/auth/login", methods=["POST"])
def api_login():
    payload = request.get_json(silent=True) or {}
    login_input = str(payload.get("login_input", "")).lower().strip()
    password = str(payload.get("password", ""))

    if not login_input or not password:
        return jsonify({"ok": False, "error": "login_input and password are required"}), 400

    if login_input == "admin" and password == "admin123":
        token = create_auth_token({
            "role": "admin",
            "admin_name": "Administrator",
        })
        return jsonify({
            "ok": True,
            "token": token,
            "role": "admin",
            "admin_name": "Administrator",
        })

    students = load_students()
    student = students.get(login_input)
    if not student or student.get("password") != password:
        return jsonify({"ok": False, "error": "Invalid credentials"}), 401

    token = create_auth_token({
        "role": "student",
        "roll_number": student["roll_number"],
        "student_name": student["name"],
        "shortname": login_input,
    })
    return jsonify({
        "ok": True,
        "token": token,
        "role": "student",
        "student_name": student["name"],
        "roll_number": student["roll_number"],
        "shortname": login_input,
    })

@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    return jsonify({"ok": True})

@app.route("/api/student/dashboard")
@api_auth_required(role="student")
def api_student_dashboard():
    roll_number = request.auth["roll_number"]
    today = datetime.now().strftime("%Y-%m-%d")
    attendance = load_attendance()

    today_attendance = {}
    for subject in CLASS_SCHEDULE:
        attendance_key = f"{roll_number}_{subject}_{today}"
        today_attendance[subject] = attendance_key in attendance

    return jsonify({
        "ok": True,
        "student_name": request.auth["student_name"],
        "roll_number": roll_number,
        "today": today,
        "today_attendance": today_attendance,
        "class_schedule": serialize_schedule(),
    })

@app.route("/api/student/class-selection")
@api_auth_required(role="student")
def api_student_class_selection():
    roll_number = request.auth["roll_number"]
    current_time = datetime.now().time().strftime("%I:%M %p")

    available_classes = {}
    for subject, schedule in CLASS_SCHEDULE.items():
        available_classes[subject] = {
            "is_available": is_class_time(subject),
            "has_marked": has_marked_attendance_today(roll_number, subject),
            "start_time": schedule["start"].strftime("%I:%M %p"),
            "end_time": schedule["end"].strftime("%I:%M %p"),
        }

    return jsonify({
        "ok": True,
        "current_time": current_time,
        "available_classes": available_classes,
    })

@app.route("/api/student/confirm-attendance", methods=["POST"])
@api_auth_required(role="student")
def api_confirm_attendance():
    payload = request.get_json(silent=True) or {}
    subject = str(payload.get("subject", "")).strip()
    if subject not in CLASS_SCHEDULE:
        return jsonify({"ok": False, "error": "Invalid subject selected"}), 400

    roll_number = request.auth["roll_number"]
    student_name = request.auth["student_name"]

    if not is_class_time(subject):
        schedule = CLASS_SCHEDULE[subject]
        return jsonify({
            "ok": False,
            "error": (
                f"Attendance for {subject} can only be marked between "
                f"{schedule['start'].strftime('%I:%M %p')} and {schedule['end'].strftime('%I:%M %p')}."
            ),
        }), 400

    if has_marked_attendance_today(roll_number, subject):
        return jsonify({
            "ok": False,
            "error": f"You have already marked attendance for {subject} today.",
        }), 409

    attendance = load_attendance()
    today = datetime.now().strftime("%Y-%m-%d")
    current_time = datetime.now().strftime("%H:%M:%S")

    attendance_key = f"{roll_number}_{subject}_{today}"
    record = {
        "roll_number": roll_number,
        "student_name": student_name,
        "subject": subject,
        "date": today,
        "time": current_time,
        "timestamp": datetime.now().isoformat(),
    }
    attendance[attendance_key] = record
    save_attendance(attendance)

    return jsonify({
        "ok": True,
        "message": f"Attendance marked successfully for {subject}.",
        "record": record,
    })

@app.route("/api/admin/summary")
@api_auth_required(role="admin")
def api_admin_summary():
    attendance = load_attendance()
    students = load_students()
    attendance_summary = build_attendance_summary(attendance)
    sorted_dates = sorted(attendance_summary.keys(), reverse=True)
    total_attendance = sum(
        len(records)
        for subjects in attendance_summary.values()
        for records in subjects.values()
    )

    return jsonify({
        "ok": True,
        "total_students": len(students),
        "total_subjects": len(CLASS_SCHEDULE),
        "total_attendance": total_attendance,
        "sorted_dates": sorted_dates,
        "attendance_summary": attendance_summary,
    })

@app.route("/api/admin/export")
@api_auth_required(role="admin")
def api_admin_export():
    attendance = load_attendance()
    wb = build_attendance_workbook(attendance)

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/')
def index():
    if 'roll_number' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_input = request.form.get('login_input', '').lower().strip()
        password = request.form.get('password')
        
        # Check for admin login
        if login_input == 'admin' and password == 'admin123':
            session['is_admin'] = True
            session['admin_name'] = 'Administrator'
            flash('Welcome, Administrator!', 'success')
            return redirect(url_for('admin'))
        
        students = load_students()
        
        # Check if login_input is a shortname
        if login_input in students and students[login_input]['password'] == password:
            session['roll_number'] = students[login_input]['roll_number']
            session['student_name'] = students[login_input]['name']
            session['shortname'] = login_input
            flash(f'Welcome, {students[login_input]["name"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid shortname or password. Please try again.', 'error')
    
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'roll_number' not in session:
        return redirect(url_for('login'))
    
    # Get today's attendance for this student
    attendance = load_attendance()
    today = datetime.now().strftime('%Y-%m-%d')
    roll_number = session['roll_number']
    
    today_attendance = {}
    for subject in CLASS_SCHEDULE.keys():
        attendance_key = f"{roll_number}_{subject}_{today}"
        today_attendance[subject] = attendance_key in attendance
    
    return render_template('dashboard.html', 
                         today_attendance=today_attendance,
                         class_schedule=CLASS_SCHEDULE)

@app.route('/class_selection')
def class_selection():
    if 'roll_number' not in session:
        return redirect(url_for('login'))
    
    # Check which classes are currently available
    available_classes = {}
    current_time = datetime.now().time()
    
    for subject, schedule in CLASS_SCHEDULE.items():
        is_available = is_class_time(subject)
        has_marked = has_marked_attendance_today(session['roll_number'], subject)
        
        available_classes[subject] = {
            'is_available': is_available,
            'has_marked': has_marked,
            'start_time': schedule['start'].strftime('%I:%M %p'),
            'end_time': schedule['end'].strftime('%I:%M %p')
        }
    
    return render_template('class_selection.html', 
                         available_classes=available_classes,
                         current_time=current_time.strftime('%I:%M %p'))

@app.route('/mark_attendance/<subject>')
def mark_attendance(subject):
    if 'roll_number' not in session:
        return redirect(url_for('login'))
    
    roll_number = session['roll_number']
    
    # Validate subject
    if subject not in CLASS_SCHEDULE:
        flash('Invalid subject selected.', 'error')
        return redirect(url_for('class_selection'))
    
    # Check if it's the right time for this class
    if not is_class_time(subject):
        schedule = CLASS_SCHEDULE[subject]
        flash(f'Attendance for {subject} can only be marked between {schedule["start"].strftime("%I:%M %p")} and {schedule["end"].strftime("%I:%M %p")}.', 'error')
        return redirect(url_for('class_selection'))
    
    # Check if already marked attendance today
    if has_marked_attendance_today(roll_number, subject):
        flash(f'You have already marked attendance for {subject} today.', 'warning')
        return redirect(url_for('dashboard'))
    
    return render_template('mark_attendance.html', subject=subject)

@app.route('/confirm_attendance/<subject>', methods=['POST'])
def confirm_attendance(subject):
    if 'roll_number' not in session:
        return redirect(url_for('login'))
    
    roll_number = session['roll_number']
    student_name = session['student_name']
    
    # Final validation
    if not is_class_time(subject) or has_marked_attendance_today(roll_number, subject):
        flash('Unable to mark attendance. Please try again.', 'error')
        return redirect(url_for('class_selection'))
    
    # Mark attendance
    attendance = load_attendance()
    today = datetime.now().strftime('%Y-%m-%d')
    current_time = datetime.now().strftime('%H:%M:%S')
    
    attendance_key = f"{roll_number}_{subject}_{today}"
    attendance[attendance_key] = {
        'roll_number': roll_number,
        'student_name': student_name,
        'subject': subject,
        'date': today,
        'time': current_time,
        'timestamp': datetime.now().isoformat()
    }
    
    save_attendance(attendance)
    
    flash(f'Attendance marked successfully for {subject}!', 'success')
    return render_template('confirmation.html', subject=subject, time=current_time)

@app.route('/admin')
def admin():
    # Check admin authentication
    if 'is_admin' not in session:
        flash('Admin access required. Please login as admin.', 'error')
        return redirect(url_for('login'))
    
    attendance = load_attendance()
    students = load_students()
    
    # Group attendance by date and subject
    attendance_summary = build_attendance_summary(attendance)
    
    # Sort by date (most recent first)
    sorted_dates = sorted(attendance_summary.keys(), reverse=True)
    
    return render_template('admin.html', 
                         attendance_summary=attendance_summary,
                         sorted_dates=sorted_dates,
                         total_students=len(students))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route('/admin/export')
def export_attendance():
    """Export attendance data to Excel file"""
    if 'is_admin' not in session:
        flash('Admin access required.', 'error')
        return redirect(url_for('login'))
    
    attendance = load_attendance()
    wb = build_attendance_workbook(attendance)
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=f'attendance_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/debug/time')
def debug_time():
    """Debug route to show current server time and class availability"""
    actual_time = datetime.now()
    simulated_time = time(9, 10)  # Same as in is_class_time function
    debug_info = {
        'actual_server_time': actual_time.strftime('%Y-%m-%d %H:%M:%S'),
        'simulated_time': simulated_time.strftime('%H:%M:%S'),
        'note': 'Using simulated time 9:10 AM for testing',
        'class_status': {}
    }
    
    for subject, schedule in CLASS_SCHEDULE.items():
        debug_info['class_status'][subject] = {
            'start': schedule['start'].strftime('%H:%M:%S'),
            'end': schedule['end'].strftime('%H:%M:%S'),
            'is_active': is_class_time(subject)
        }
    
    return f"<pre>{json.dumps(debug_info, indent=2)}</pre>"

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
